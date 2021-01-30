# words = ["cat", "dog", "tiger", "jamal"]
#
# words.append("Ladybug")
# words.remove("cat")
# mr_jones = words.index("dog")
# # words.sort(reverse = True)
# words.insert(0, "Donkey")
# words.insert(0, "Donk")
#
# words.sort(key = str.lower)
# print(words)
# # print(mr_jones)

# output = ''
# for i in range(4):
#     output += str(i) + " "
# print(output)
# lyric = "Cause it's always a good time!"
# office = ["We", "don't", "even", "have", "to", "try!"]
#
# for i in range(len(office)):
#     print(office[i] + " ")
# print(lyric)
# def emoji_converter(mess):
#     message = msg.split()
#     emoji = {
#         "XD": "😂️",
#         "<3": "😍",
#         ":)": "😊",
#         ":p": "😜",
#         "gg": "+5 karma",
#         "Elliott": "baby"
# }
#     output = ''
#     for i in message:
#         output += emoji.get(i, i) + " "
#     print(output)

#
# msg = input("Ian: ")
# emoji_converter(msg)
#
# office = ["staples", "paper", "desk", "monitor"]
#
# for word in range(len(office)):
#     print("Index " + str(word) + " and the item: " + office[word])

# Simpsons = ["Doh!", "Apu", "Krusty the Clown"]
# phrase, character, character1 = Simpsons
# star = ["rush"]
# name = "Larry"
# clothes = "Orange suit"
# name, clothes = clothes, name
# print("Name   : " + name)
# print("Clothes: " + clothes)

# print(range(len(Simpsons)))
# for i in range(len(Simpsons)):
#     print(Simpsons[i])

# george = ["with", "the", "yellow", "hat"]
# print(42 not in george)
# print("yellow" in george)
# print("man" in george)

# the_cat = "in the hat."
# committee = "zzz" in the_cat
# print(committee)
# name = "Bob"
# for char in name:
#     print(char)
#     #NOTE: strings in python are immutable
# state = "The witness could not have known what time it was unless..."
#
# print(state[0:18] + "most defintely" + state[21:59])

#
# import copy
# scroll = ["letter", "pen", "ink"]
# envelope = copy.deepcopy(scroll)
# envelope[1] = 34
# print(scroll)
# print(envelope)

# saying = "Pegasus! Rest assured, I will save my grandpa, my" \
#          " friends and Yugi; so that I can put an end to this" \
#         " CHARADE."
# awkward = r'That\'s not our son!'
# print(awkward)
# statement = """Hello there I am am na. A man with a plan who likes
# to know how you are  making that ice cream sandwich with the barbel
# and the jackel. You are kind of a more on that later, let's talk
# about me. Me and my big book of words."""
# print(statement)
# print(range(len(statement)))
#
# import utils
# numbers = [3, 1, 200, 2, 5, 100]
# print(numbers)
# maximum = utils.find_max(numbers)
# print(maximum)
# import random
# from utils import party
# class DiceFunctions:
#     def roll(self):
#         first = random.randint(1, 6)
#         second = random.randint(1, 6)
#         return first, second
#
#
# diceRoll = DiceFunctions()
# print(diceRoll.roll())
#
# pokemon = ["Squirtle", "Charizar", "Pikachu"]
# nintendo = party(pokemon)
#
# print(f"{nintendo} is unable to battle.")

# class Person:
#     def __init__(self, name, age, height):
#         self.name = name
#         self.age = age
#         self.height = height
#
#     def print(self):
#         print(f"Name: {self.name}")
#         print(f"Age: {self.age}")
#         print(f"Height: {self.height}")
#
#
# class Woman(Person):
#     pass
#
#
# class Man(Person):
#     pass
#
#
# Bob = Man("Bob", 28, "5'11")
# Bob.print()
# Samantha = Woman("Samantha", 26, "5'9")
# Samantha.print()

# list = [92, "Dave", 2.34]
# list.append("49'ers")
# list1 = [list + "yelp"]
# for i in list:
#     print(i)

#
# class Parent:
#     def hello_func(self):
#         print("Hey buddy")
#
#
# class Child(Parent):
#     def hello_func(self): #Function overriding
#         print("Hello Mr./Ms.!")
#         super().hello_func()
#
#
# greeting = Child()
#
# greeting.hello_func()

# class A:
#     def __init__(self, a):
#         self.a = a
#
#     def __add__(self, other):
#         return self.a + other.a
#
#     def __mul__(self, other):
#         return self.a * other.a
#
#     def __sub__(self, other):
#         return self.a - other.a
#
#
# a = int(input("Obj1: "))
# b = int(input("Obj2: "))
# print(f"{a} * {b}")
# obj1 = A(a)
# obj2 = A(b)
# print(obj1 * obj2)
# print(f"{a} - {b}")
# print(obj1 - obj2)
#
#
# class Car:
#     num_wheels = 4
#
#
# BMW = Car()
# print(BMW.num_wheels)
#
# from pathlib import Path
#
# path = Path("emails")
# print(path.rmdir())

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(1, 1)
# print(cell.value)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    # print(cell.value)
    corrected_price = cell.value * 0.90
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
    print(round(corrected_price_cell.value, 2))
wb.save('transactions1.xslx')
values = Reference(sheet,
          min_row = 2,
          max_row = sheet.max_row,
          min_col = 4,
          max_col = 4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')